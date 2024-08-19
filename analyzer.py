import os
import re
import pandas as pd
import openpyxl as xl
from datetime import datetime
from tkinter import messagebox
from einqueryer import EinQueryer


class Analyzer:    
    """
    This class is responsible for analyzing a visit day of a collector. When requested to make
    an analysis, it will require a set of EINs (the visited establishments), the collector and
    the date the visit was made. It will first check if there is an entry for the queried EIN,
    if so, it'll check if the collector of the entry match with the given analyzed collector.
    If there's a mismatch between the analyzed collector and the visit entry's collector, the
    analyzer will update the entry to have the analyzed collector instead. If there is no entry
    with the analyzed EIN, the Analyzer will request information about the EIN for a EIN query
    service. Uppon success, it will insert the retrieved EIN info as an visit entry for the
    analyzed collector and date. 
    If no query was made (the establishment was found in the database), it'll check if the
    found visit entry has as a visit day the weekday of the analyzed date. If not, the Analyzer
    will append the analyzed date weekday to the visit days list of the visit entry. After it,
    the EIN will be marked as visited. All EINs that doesn't match will be marked as non-visited
    and can be requested via a property.
    """


    DB_FILE = os.path.expandvars(r'%USERPROFILE%\OneDrive\Nota Paraná\Estabelecimentos.xlsx')


    def __init__(self):
        self.df = pd.read_excel(Analyzer.DB_FILE)
        self.last_row = len(self.df) + 2
        self.df.drop_duplicates(subset=['CNPJ'], inplace=True)
        self.wb = xl.load_workbook(Analyzer.DB_FILE)
        self.ws = self.wb.active
    

    def get_nonvisited(self, eins: set[str], collector: str, date: datetime):
        for ein in eins:
            ein = ein.replace('.', '')
            f = self.df[self.df['CNPJ'] == ein]

            if f.empty:
                self.query_ein(ein, collector, date)
            else:
                self.df = self.df[self.df['Coletor'] == collector] # Extract only the entries of the analyzed collector
                self.df.dropna(inplace=True)

                if f.iloc[0]['Coletor'] != collector:
                    self.update_collector(f, collector)

                if not str(date.weekday() + 2) in str(f.iloc[0]['Dias']):
                    self.append_visit_day(f, date)
                    
        self.df = self.df[self.df['Dias'].str.contains(str(date.weekday() + 2), na=False)]
        return self.df.loc[~self.df['CNPJ'].isin(eins)]


    def query_ein(self, ein: str, collector: str, date: datetime):
        establishment = EinQueryer.query(ein)

        if establishment is None:
            messagebox.showerror('Falha na Consulta', f'O CNPJ "{ein}" não pôde ser consultado')
            return
        
        self.append_establishment(establishment, collector, date)

    
    def append_establishment(self, establishment: dict, collector: str, date: datetime):
        row = self.ws[self.last_row]
        row[0].value = establishment['region']
        row[1].value = establishment['name']
        row[2].value = establishment['ein']
        row[3].value = collector
        row[4].value = str(date.weekday() + 2)
        row[5].value = establishment['address']
        row_list = [cell.value for cell in row]
        self.df.loc[len(self.df.index)] = row_list
        self.wb.save(Analyzer.DB_FILE)
    

    def append_weekday(self, days: str, date: datetime) -> list[str]:
        days = re.sub(r'\D', '', days)

        if days == '':
            return str(date.weekday() + 2)

        days = [int(d) for d in days.split(',')]
        days.append(date.weekday() + 2)
        days.sort()
        days = ','.join(str(d) for d in days)
        return days


    def update_collector(self, f: pd.DataFrame, collector: str):
        self.df.loc[f.index.item(), 'Coletor'] = collector
        row = f.index.item() + 2
        cell = self.ws.cell(row=row, column=4)
        cell.value = collector
        self.wb.save(Analyzer.DB_FILE)


    def append_visit_day(self, f: pd.DataFrame, date: datetime):
        days = self.append_weekday(str(f.iloc[0]['Dias']), date)
        row = f.index.item() + 2
        cell = self.ws.cell(row=row, column=5)
        cell.value = days
        self.wb.save(Analyzer.DB_FILE)