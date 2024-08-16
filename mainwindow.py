import os
import re
import pandas as pd
import openpyxl as xl
from datetime import datetime
from tkinter import *
from tkinter.ttk import *
from functools import partial
from analyzer import Analyzer
from tkinter import messagebox


REPORTS_PATH = r'C:\Users\j3392\OneDrive\Nota Paraná\Estabelecimentos Não Visitados\2024\Agosto'
COLLECTORS = ('Claudinei', 'Moto', 'Motorista', 'Noturno', 'Vitor')


class MainWindow(Tk):
    def __init__(self):
        super().__init__()
        self.title('Inspetor de Visitas')
        self.geometry('400x300')
        self.config(padx=5, pady=5)
        self.eins: set[str] = set()
        self.setup_collector_frame(COLLECTORS)
        self.setup_date_frame()
        self.setup_ein_frame()
        self.listbox = Listbox(self)
        self.analyzebtn = Button(self, text='Analizar', command=partial(self.analyze))
        self.listbox.pack(fill=BOTH, expand=True, pady=(5, 0))
        self.analyzebtn.pack(fill=X)
        self.ein_entry.focus()
        self.mainloop()


    def setup_collector_frame(self, collectors: list[str]):
        self.collector_var = StringVar(self, collectors[0])
        self.collector_frame = Frame(self)
        self.collector_label = Label(self.collector_frame, text='Coletor: ')
        self.collector_combo = Combobox(self.collector_frame, values=collectors, textvariable=self.collector_var)
        self.collector_label.pack(side=LEFT, anchor=N)
        self.collector_combo.pack(anchor=N, fill=X, expand=True)
        self.collector_frame.pack(fill=X, pady=(0, 5))


    def setup_date_frame(self):
        self.date_var = StringVar(self, datetime.now().strftime('%d/%m/%Y'))
        self.date_frame = Frame(self)
        self.date_label = Label(self.date_frame, text='Data: ')
        self.date_entry = Entry(self.date_frame, textvariable=self.date_var)
        self.date_label.pack(side=LEFT)
        self.date_entry.pack(side=RIGHT, fill=X, expand=True)
        self.date_frame.pack(fill=X, pady=(0, 5))
        

    def setup_ein_frame(self):
        self.ein_var = StringVar(self)
        self.ein_frame = Frame(self)
        self.ein_label = Label(self.ein_frame, text='CNPJ: ')
        self.ein_entry = Entry(self.ein_frame, textvariable=self.ein_var)
        self.ein_entry.bind('<Return>', partial(self.on_accesskey_return))
        self.ein_label.pack(side=LEFT, anchor=N)
        self.ein_entry.pack(anchor=N, fill=X, expand=True)
        self.ein_frame.pack(fill=X)


    def on_accesskey_return(self, evt):
        ein = MainWindow.extract_ein(self.ein_var.get())
        
        if ein and ein not in self.eins:
            self.eins.add(ein)
            self.listbox.insert(END, ein)
        
        self.ein_var.set('')
        self.ein_entry.focus()


    @staticmethod
    def extract_ein(text: str) -> str:
        if text.startswith('http') and 'fazenda' in text:
            match = re.search(r'p=(\d{44})', text)
            ak = match.group(1)
            return f'{ak[6:8]}.{ak[8:11]}.{ak[11:14]}/{ak[14:18]}-{ak[18:20]}'
        else:
            text = re.sub(r'\D', '', text)
            return f'{text[:2]}.{text[2:5]}.{text[5:8]}/{text[8:12]}-{text[12:14]}'


    def analyze(self):
        if not messagebox.askyesno('Revisão', f'O Coletor: {self.collector_var.get()}\ne a Data: {self.date_var.get()}\nEstão corretos?'):
            return
        
        self.config(cursor='watch')
        self.analyzebtn.config(state=DISABLED)
        analyzer = Analyzer()
        nonvisited = analyzer.get_nonvisited(self.eins, self.collector_var.get(), datetime.strptime(self.date_var.get(), '%d/%m/%Y'))
        self.save_report(nonvisited)
        self.config(cursor='arrow')


    def save_report(self, df: pd.DataFrame):
        df.reset_index(drop=True, inplace=True)
        path, wb = self.get_workbook()
        ws = wb.active

        for i, row in df.iterrows():
            for col, val in enumerate(row.to_list()):
                cell = ws.cell(i + 1, col + 1)
                cell.value = val

        wb.save(path)
        messagebox.showinfo('Relatório', 'Relatório concluído!')
    

    def get_workbook(self) -> tuple[str, xl.Workbook]:
        """
        Checks if a workbook for the current collector already exists.
        If so, loads and return it, otherwise, creates a new one.
        It also creates a worksheet for the given date.
        """
        report_path = fr'{REPORTS_PATH}\{self.collector_var.get()}.xlsx'

        if os.path.exists(report_path):
            wb = xl.load_workbook(report_path)
        else:
            wb = xl.Workbook()
            wb.remove(wb.active)

        ws_title = self.date_var.get().replace('/', '-')
        wb.create_sheet(ws_title)
        wb.active = wb[ws_title]
        return report_path, wb