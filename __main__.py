import re
import os
import requests
import pandas as pd
import openpyxl as xl
from datetime import datetime


ESTABLISHMENTS_DB = os.path.expandvars(r'%USERPROFILE%\OneDrive\Nota Paraná\Estabelecimentos.xlsx')
REPORTS_PATH = os.path.expandvars(r'%USERPROFILE%\OneDrive\Nota Paraná\Estabelecimentos Não Visitados\2024\Setembro')


def cls():
    if os.name == 'nt':
        os.system('cls')
    else:
        os.system('clear')


def get_collector(df: pd.DataFrame, eins: set[str]) -> str:
    """Figure out the collector based on the EINs passed."""
    eins = map(lambda x: x.replace('.', ''), eins)
    f = df.loc[df['CNPJ'].isin(eins)]
    f = f.drop_duplicates(subset=['CNPJ'])
    assert not f.empty, 'O coletor não pôde ser encontrado' # if this execute, we're fucked...
    f = f['Coletor'].value_counts()
    return f[f == f.max()].index[0]


def get_date() -> datetime:
    while True:
        cls()
        print(f'Informe a data (deixe vazio para "{datetime.now().strftime('%d/%m/%Y')}"):')
        d = input('>>> ')
        
        if d == '':
            return datetime.now()
        
        try:
            return datetime.strptime(d, '%d/%m/%Y')
        except Exception:
            continue


def extract_ein(text: str) -> str:
    """Extract the EIN from an QRCode URL"""
    match = re.search(r'p=(\d{44})', text)
    ak = match.group(1)
    return f'{ak[6:8]}.{ak[8:11]}.{ak[11:14]}/{ak[14:18]}-{ak[18:20]}'


def get_eins() -> set[str]:
    eins = set()
    while True:
        cls()

        for ein in eins:
            print(ein)
        
        ein = input('CNPJ: ')
        
        if ein == '':
            break

        if ein.startswith('http') and 'fazenda' in ein:
            ein = extract_ein(ein)

        eins.add(ein)
    return eins


def query_ein(ein: str) -> dict | None:
    ein = re.sub(r'\D', '', ein)
    res = requests.get('https://api-publica.speedio.com.br/buscarcnpj', {'cnpj': ein})
    
    if not res.ok:
        return None
    
    return res.json()


def update_collector(ein: str, collector: str) -> None:
    """Updates the collector column in the database file on the row of f"""
    wb = xl.load_workbook(ESTABLISHMENTS_DB)
    ws = wb.active

    for row in ws.rows:
        if row[2].value == ein:
            row[3].value = collector
            wb.save(ESTABLISHMENTS_DB)
            return


def update_days(ein: str, day: int) -> None:
    """Update the days column in the database file"""
    wb = xl.load_workbook(ESTABLISHMENTS_DB)
    ws = wb.active

    for row in ws.rows:
        if row[2].value == ein:
            days = re.sub(r'\D', '', str(row[4].value))
            days = [int(c) for c in days] # TODO: Handle Exception
            days.append(day)
            days.sort()
            days = [str(n) for n in days]
            row[4].value = ','.join(days)
            wb.save(ESTABLISHMENTS_DB)
            return


def append_establishment(establishment: dict, ein: str, collector: str, date: datetime) -> None:
    wb = xl.load_workbook(ESTABLISHMENTS_DB)
    ws = wb.active
    ws.cell(ws.max_row+1, 1).value = establishment['BAIRRO'] if 'BAIRRO' in establishment else '???'
    
    if 'RAZAO SOCIAL' in establishment and establishment['RAZAO SOCIAL'] != '':
        name = establishment['RAZAO SOCIAL']
    elif 'NOME FANTASIA' in establishment and establishment['NOME FANTASIA'] != '':
        name = establishment['NOME FANTASIA']
    else:
        name = '???'
    
    ws.cell(ws.max_row+1, 2).value = name
    ws.cell(ws.max_row+1, 3).value = ein
    ws.cell(ws.max_row+1, 4).value = collector
    ws.cell(ws.max_row+1, 5).value = date.weekday() + 2

    if 'LOGRADOURO' in establishment and 'NUMERO' in establishment:
        address, number = establishment['LOGRADOURO'], establishment['NUMERO']
    else:
        address, number = '???', '???'

    ws.cell(ws.max_row+1, 6).value = f'{address}, {number}'


def find_and_update(df: pd.DataFrame, collector: str, date: datetime, eins: set[str]) -> None:
    for ein in eins:
        ein = ein.replace('.', '')
        f = df[df['CNPJ'] == ein]

        if f.empty:
            print(f'Consultando o CNPJ "{ein}"')
            establishment = query_ein(ein)

            if establishment is not None:
                append_establishment(establishment, ein, collector, date)
        else:
            if f.iloc[0]['Coletor'] != collector:
                print(f'Atualizando o Coletor para o CNPJ "{ein}"')
                update_collector(ein, collector)

            if str(date.weekday() + 2) not in str(f.iloc[0]['Dias']):
                print(f'Atualizando os Dias de Visita para o CNPJ "{ein}"')
                update_days(ein, date.weekday() + 2)


def get_workbook(collector: str, date: datetime) -> tuple[str, xl.Workbook]:
    """
    Checks if a workbook for the current collector already exists.
    If so, loads and return it, otherwise, creates a new one.
    It also creates a worksheet for the given date.
    """
    report_path = fr'{REPORTS_PATH}\{collector}.xlsx'

    if os.path.exists(report_path):
        wb = xl.load_workbook(report_path)
    else:
        wb = xl.Workbook()
        wb.remove(wb.active)

    ws_title = date.strftime('%d-%m')
    wb.create_sheet(ws_title)
    wb.active = wb[ws_title]
    return report_path, wb


def save_nonvisited(df: pd.DataFrame, collector: str, date: datetime, eins: set[str]) -> None:
    print(f'Salvando relatório de estabelecimentos não visitados para o coletor "{collector}"')
    df = df[df['Coletor'] == collector]
    df = df[df['Dias'].str.contains(str(date.weekday() + 2), na=False)]
    path, wb = get_workbook(collector, date)
    ws = wb.active
    eins = [e.replace('.', '') for e in eins]
    nonvisited = df[~df['CNPJ'].isin(eins)]
    r = 1
    
    for row in nonvisited.iterrows():
        row = row[1] # Skip the index, get the Series
        ws.cell(r, 1).value = row['Região']
        ws.cell(r, 2).value = row['Estabelecimento']
        ws.cell(r, 3).value = row['CNPJ']
        ws.cell(r, 4).value = row['Dias']
        ws.cell(r, 5).value = row['Endereço']
        r += 1
    
    wb.save(path)


def main():
    df = pd.read_excel(ESTABLISHMENTS_DB, dtype={'CNPJ': str, 'Dias': str})
    date = get_date()
    eins = get_eins()
    collector = get_collector(df, eins)
    cls()
    find_and_update(df, collector, date, eins)
    save_nonvisited(df, collector, date, eins)
    print('Análize Finalizada')


if __name__ == '__main__':
    main()